from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import load_workbook
from selenium.webdriver.common.alert import Alert

wb = load_workbook("/home/bandaru/Desktop/Selenium/RPA_task.xlsx")
print(wb)
sheet = wb['Sheet1']
headers = [cell.value for cell in sheet[1]]
print(headers)
table_data = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
    data = {headers[i]: row[i].value for i in range(len(headers))}
    table_data.append(data)
print(table_data)
driver = webdriver.Chrome()
driver.get("https://rpachallenge.com/")
driver.maximize_window()
alert = Alert(driver)
driver.implicitly_wait(10)
k = 2
for j in table_data:
    clear = driver.find_element(By.CSS_SELECTOR, 'div:nth-child(1) > rpa1-field > div > input')
    clear.click()

    F_Name = driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelFirstName"]')
    F_Name.send_keys(str(j['First Name']))

    L_Name = driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelLastName"]')
    L_Name.send_keys(str(j['Last Name']))

    company_Name = driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelCompanyName"]')
    company_Name.send_keys(str(j['Company Name']))

    email = driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelEmail"]')
    email.send_keys(str(j['Email']))

    role = driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelRole"]')
    role.send_keys(str(j['Role in Company']))

    PhNo = driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelPhone"]')
    PhNo.send_keys(str(j['Phone Number']))

    Address = driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelAddress"]')
    Address.send_keys(str(j['Address']))

    driver.implicitly_wait(10)
    Submit = driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]')
    Submit.click()
    k += 1

wb.save("/home/bandaru/Desktop/Selenium/RPA_task.xlsx")
driver.quit()