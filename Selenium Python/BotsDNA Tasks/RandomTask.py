from openpyxl import load_workbook
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
wb=load_workbook("/home/bandaru/Downloads/RPA_task (1).xlsx")
print(wb)

sheet=wb['Sheet1']
print(sheet)
headers=[cell.value for cell in sheet[1]]
print(headers)
table_data=[]

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,max_col=sheet.max_column):
    data={headers[i]:row[i].value for i in range(len(row))}
    table_data.append(data)
print(table_data)
driver=webdriver.Chrome()
driver.get("https://rpachallenge.com/")
driver.maximize_window()
time.sleep(4)
for i in table_data:
    wait = WebDriverWait(driver, 5)
    fname = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,'input[ng-reflect-name="labelFirstName"]')))
    fname.click()
    fname.send_keys(i['First Name'])
    lname = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-reflect-name="labelLastName"]')))
    lname.click()
    lname.send_keys(i['Last Name'])
    cname = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-reflect-name="labelCompanyName"]')))
    cname.click()
    cname.send_keys(i['Company Name'])
    rname = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-reflect-name="labelRole"]')))
    rname.click()
    rname.send_keys(i['Role in Company'])
    email = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-reflect-name="labelEmail"]')))
    email.click()
    email.send_keys(i['Email'])
    phone = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-reflect-name="labelPhone"]')))

    phone.click()
    phone.send_keys(i['Phone Number'])
    add = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-reflect-name="labelAddress"]')))

    add.click()
    fname.send_keys(i['Address'])
    bu = driver.find_element(By.XPATH, '/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/input')
    bu.click()








    # wait = WebDriverWait(driver, 5)
    # email = wait.until(EC.presence_of_element_located(
    #     (By.XPATH, f"/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[{'*'}]/rpa1-field/div/input")))
    # email.click()
    # email.send_keys("i['Last Name']")
    # wait = WebDriverWait(driver, 5)
    # email = wait.until(EC.presence_of_element_located(
    #     (By.XPATH, f"/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[{'*'}]/rpa1-field/div/input")))
    # email.click()
    # email.send_keys("i['Company Name']")
    # wait = WebDriverWait(driver, 5)
    # email = wait.until(EC.presence_of_element_located(
    #     (By.XPATH, f"/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[{'*'}]/rpa1-field/div/input")))
    # email.click()
    # email.send_keys("i['Role in Company']")
    # wait = WebDriverWait(driver, 5)
    # email = wait.until(EC.presence_of_element_located(
    #     (By.XPATH, f"/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[{'*'}]/rpa1-field/div/input")))
    # email.click()
    # email.send_keys("i['Email']")
    # wait = WebDriverWait(driver, 5)
    # email = wait.until(EC.presence_of_element_located(
    #     (By.XPATH, f"/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[{'*'}]/rpa1-field/div/input")))
    # email.click()
    # email.send_keys("i['Phone Number']")
    # wait = WebDriverWait(driver, 5)
    # email = wait.until(EC.presence_of_element_located(
    #     (By.XPATH, f"/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[{'*'}]/rpa1-field/div/input")))
    # email.click()
    # email.send_keys("i['Address']")


#/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[3]/rpa1-field/div/input
#/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[4]/rpa1-field/div/input





#/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[1]/rpa1-field/div/input
#/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/div/div[5]/rpa1-field/div/input

# try:
#     wait = WebDriverWait(driver, 5)
#     wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[type="button" i]')))
#     print("Button indentified")
# except:
#     print("Button not indentified.")
el=driver.find_elements(By.CSS_SELECTOR,'input[name="Buyer"]')