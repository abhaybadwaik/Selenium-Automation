from openpyxl import load_workbook
import time
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver

wb=load_workbook("/home/bandaru/Desktop/Selenium/Currency_Convertor.xlsx")
ws=wb["Sheet1"]
headers=[i.value for i in ws[1]]
print(headers)

l1=[]
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=ws.max_column):
    res={headers[i]:row[i].value for i in range(len(headers))}
    l1.append(res)
print(l1)

driver=webdriver.Chrome()
driver.get("https://www.xe.com/currencyconverter/")
driver.maximize_window()
time.sleep(5)
j=2
for i in l1:
    amount = driver.find_element(By.XPATH,'//*[@id="amount"]')
    amount.click()
    amount.send_keys(Keys.CONTROL+'A')
    amount.send_keys(i["Amount"])

    from1=driver.find_element(By.XPATH,'//*[@id="midmarketFromCurrency"]/div[2]/div/input')
    from1.click()
    from1.send_keys(i["From"])
    time.sleep(3)
    from1.send_keys(Keys.ENTER)

    to1=driver.find_element(By.XPATH,'//*[@id="midmarketToCurrency"]/div[2]/div/input')
    to1.click()
    to1.send_keys(i["To"])
    time.sleep(2)
    to1.send_keys(Keys.ENTER)
    time.sleep(2)

    convert=driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[5]/div[2]/div[1]/div[1]/div/div[2]/div[3]/button')
    convert.click()
    time.sleep(3)

    result=driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[5]/div[2]/div[1]/div[1]/div/div[2]/div[3]/div/div[1]/div[1]/p[2]')
    value1=result.text
    print(value1)
    ws["D"+str(j)] = value1
    j+=1

    convert1=driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[5]/div[2]/div[1]/div[1]/div/div[2]/div[1]/a[1]')
    convert1.click()
    time.sleep(3)
    wb.save("/home/bandaru/Desktop/Selenium/Currency_Convertor.xlsx")

