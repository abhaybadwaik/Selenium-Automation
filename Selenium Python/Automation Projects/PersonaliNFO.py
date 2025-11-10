from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import load_workbook,Workbook
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import Select
from datetime import datetime

wb=load_workbook("/home/bandaru/Desktop/Selenium/fruit.xlsx")
sheet=wb['Sheet1']
headers=[cell.value for cell in sheet[1]]
print(headers)
table_data=[]
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,max_col=sheet.max_column):
    data={headers[i]:row[i].value for i in range(len(headers))}
    table_data.append(data)
print(table_data)
driver=webdriver.Chrome()
driver.get("file:///home/bandaru/Desktop/Selenium/Basics.html")
driver.maximize_window()
alert = Alert(driver)
time.sleep(3)
for j in table_data:
    email= driver.find_element(By.CSS_SELECTOR, 'input[type="email"]')
    email.send_keys(j['email'])
    email.clear()
    time.sleep(1)

    passwd= driver.find_element(By.CSS_SELECTOR, 'input[type="password"]')
    passwd.send_keys(j['password'])
    passwd.clear()
    time.sleep(1)

    if j['gender']=='F':
        gender= driver.find_element(By.CSS_SELECTOR, 'input[value="female"]')
        gender.click()
        time.sleep(1)
    elif j['gender']=='M':
        gender = driver.find_element(By.CSS_SELECTOR, '[type="radio"]')
        gender.click()
        time.sleep(1)

    else:
        gender=driver.find_element(By.CSS_SELECTOR,'[value="others"]')
        gender.click()

    fruit = driver.find_element(By.ID, 'fruits')
    fruit.click()
    select = Select(fruit)
    select.select_by_visible_text(j['Fruit'])
    time.sleep(1)


    checkbox = driver.find_element(By.CSS_SELECTOR, '[type="checkbox"]')
    checkbox.click()
    time.sleep(1)

    DOB = driver.find_element(By.CSS_SELECTOR, '[type="date"]')
    DOB.click()

    dob_value = j['DOB']  # This is a datetime object

    # Convert datetime to string format "DD/MM/YY"
    dob_value_str = dob_value.strftime("%d/%m/%Y")  # Example: "08/01/2003"

    # Split into day, month, and year
    dob_parts = dob_value_str.split("/")  # ["08", "01", "2003"]
    day = dob_parts[0]   # "08"
    month = dob_parts[1] # "01"
    year = dob_parts[2]

    DOB.send_keys(day)   # Enter "08"
    DOB.send_keys(month)  # Enter "01"
    DOB.send_keys(year)   # Enter "2003"time.sleep(2)

    time.sleep(2)

    Address= driver.find_element(By.ID, 'address')
    Address.click()
    Address.clear()
    Address.send_keys(j['Address'])

    Submit= driver.find_element(By.CSS_SELECTOR, '[value="Submit"]')
    Submit.click()
    time.sleep(1)

    # print(alert.text)
    # alert.accept()



wb.save("/home/bandaru/Desktop/Selenium/fruit.xlsx")
driver.quit()