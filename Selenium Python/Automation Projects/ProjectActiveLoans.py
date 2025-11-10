from zipfile import ZipFile
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

wb=load_workbook("/home/bandaru/Desktop/Selenium/input.xlsx")
ws=wb.active
headers=[i.value for i in ws[1]]
print(headers)
dict=[]
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=ws.max_row):
    res={headers[i]:row[i].value for i in range(len(headers))}
    dict.append(res)
print(dict)

driver=webdriver.Chrome()
driver.get("https://botsdna.com/ActiveLoans/")
time.sleep(3)
driver.maximize_window()
for i in dict:
    zipfile=driver.find_element(By.CSS_SELECTOR,f"a[href='{i['AccountNumber']}.zip']")
    zipfile.click()
    driver.implicitly_wait(10)

for j in dict:
    with ZipFile(f"/home/bandaru/Downloads/{j['AccountNumber']}.zip", "r") as unzip:
            unzip.extractall("/home/bandaru/Downloads/ABHAY")

el=driver.find_element(By.TAG_NAME,"table")
table_html =el.get_attribute('outerHTML')
soup = BeautifulSoup(table_html, 'html.parser')
rows = soup.find_all('tr')
headings=[th.get_text() for th in rows[0].find_all('td')]
print(headings)
data=[]
for row in rows[1:]:
    value1=[th.get_text() for th in row.find_all('td')]
    cells={headings[i]:value1[i] for i in range(len(headings))}
    data.append(cells)
print(data)

m=2
for j in data:
    for s in dict:
        s1=s['AccountNumber']
        s2=s1[8:]
        print(s2)
        if j['LOAN CODE'].endswith(s2):
            ws['H' + str(m)] = j['STATUS']
            ws['G' + str(m)] = j['PAN NUMBER']
            m+=1

k=2
for z in dict:
    txt = open(f"/home/bandaru/Downloads/ABHAY/{z['AccountNumber']}.txt")
    t1=txt.read()
    # print(t1)
    f=t1.split("\n")
    # print(f)
    list2=[]
    for i in f:
        if i!='':
            t=i.split(":")
            # print(t)
            val=t[1].strip()
            list2.append(val)
    ws["B"+str(k)]=list2[1]
    ws["C"+str(k)]=list2[2]
    ws["D"+str(k)]=list2[3]
    ws["E"+str(k)]=list2[4]
    ws["F"+str(k)]=list2[5]
    k+=1
    print(list2)

wb.save("/home/bandaru/Desktop/Selenium/input.xlsx")