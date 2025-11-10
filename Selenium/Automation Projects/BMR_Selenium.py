from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook,Workbook
wb=load_workbook('/home/bandaru/Desktop/Selenium/BMR_Data.xlsx')
sheet=wb['Sheet1']
headers=[cell.value for cell in sheet[1]]
print(headers)
table_data=[]
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,max_col=sheet.max_column):
    data={headers[i]:row[i].value for i in range(len(headers))}
    table_data.append(data)
print(table_data)
driver=webdriver.Chrome()
driver.get("https://www.calculator.net/bmr-calculator.html")
driver.maximize_window()
time.sleep(1)
k=2
for j in table_data:
    z = driver.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div[2]/form/table[4]/tbody/tr[2]/td/input[3]') #CLEAR TAb hai
    if 15<=j['Age']<=80:
        age = driver.find_element(By.XPATH,'/html/body/div[3]/div[1]/div[2]/div[2]/form/table[1]/tbody/tr[1]/td[2]/input')
        age.send_keys(str(j['Age']))
        if j['Gender'].lower()=='male':
            m=driver.find_element(By.XPATH,'/html/body/div[3]/div[1]/div[3]/div[2]/form/table[1]/tbody/tr[2]/td[2]/label[1]')
            m.click()
        elif j['Gender'].lower()=='female':
            f=driver.find_element(By.XPATH,'/html/body/div[3]/div[1]/div[2]/div[2]/form/table[1]/tbody/tr[2]/td[2]/label[2]')
            f.click()
        height=driver.find_element(By.XPATH,"/html/body/div[3]/div[1]/div[3]/div[2]/form/table[3]/tbody/tr[1]/td[2]/input")
        height.send_keys(str(j['Height']))
        weight=driver.find_element(By.XPATH,'/html/body/div[3]/div[1]/div[3]/div[2]/form/table[3]/tbody/tr[2]/td[2]/input')
        weight.send_keys(str(j['Weight']))
        calculate = driver.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div[2]/form/table[4]/tbody/tr[2]/td/input[2]')
        calculate.submit()
        result=driver.find_element(By.XPATH,'/html/body/div[3]/div[1]/div[3]/div[1]/font/b')
        out= result.text
        print(out)
        j['Result']=out
        sheet['G'+str(k)]=out
        k+=1
    else:
        j['Result']='Invalid'
        sheet['G' + str(k)] ='Invalid'
        k+=1
wb.save("/home/bandaru/Desktop/Selenium/BMR_Data.xlsx")
driver.quit()