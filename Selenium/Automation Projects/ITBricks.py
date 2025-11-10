from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import load_workbook,Workbook
from selenium.webdriver.common.alert import Alert

wb=load_workbook("/home/bandaru/Desktop/Selenium/holidays.xlsx")
sheet=wb['Sheet1']
headers=[cell.value for cell in sheet[1]]
print(headers)
table_data=[]
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,max_col=sheet.max_column):
    data={headers[i]:row[i].value for i in range(len(headers))}
    table_data.append(data)
print(table_data)
driver=webdriver.Chrome()
driver.get("https://rpademo.automationanywhere.com/itbricks_crm.php")
driver.maximize_window()
alert = Alert(driver)
time.sleep(2)
k=2
for j in table_data:
    clear= driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[7]/td/input[2]')
    clear.click()
    F_Name = driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[1]/td/input')
    if j['First Name']==None:
        F_Name.clear()
    else:
        F_Name.send_keys(str(j['First Name']))

    L_Name = driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[2]/td/input')
    if j['Last Name'] == None:
        L_Name.clear()
    else:
        L_Name.send_keys(str(j['Last Name']))

    company_Name = driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[3]/td/input')
    if j['Company'] == None:
        company_Name.clear()
    else:
        company_Name.send_keys(str(j['Company']))

    email = driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[4]/td/input')
    if j['Email']==None:
        email.clear()
    else:
        email.send_keys(str(j['Email']))

    No = driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[5]/td/input')
    if j['Phone']==None:
        No.clear()
    else:
        No.send_keys(str(j['Phone']))

    Country = driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[6]/td/input')
    if j['Country']==None:
        Country.clear()
    else:
        Country.send_keys(str(j['Country']))

    Reg = driver.find_element(By.XPATH, '/html/body/form/center/table/tbody/tr[7]/td/input[1]')
    Reg.click()
    try:
        a=alert.text
        print(a)
        alert.accept()
        sheet['G' + str(k)] = a
    except:
        print("Registerd")
        sheet['G'+str(k)]="Registration"
    k+=1

wb.save("/home/bandaru/Desktop/Selenium/holidays.xlsx")
driver.quit()
