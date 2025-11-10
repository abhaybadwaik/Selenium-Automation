from openpyxl import load_workbook
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

wb=load_workbook("/home/bandaru/Desktop/Selenium/input.xlsx")
sheet=wb['Sheet1']
headers=[cell.value for cell in sheet[1]]
print(headers)
table_data=[]
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,max_col=sheet.max_column):
    data={headers[i]:row[i].value for i in range(len(row))}
    table_data.append(data)
print(table_data)

driver=webdriver.Chrome()
driver.get("https://botsdna.com/vitrualplots/")
driver.maximize_window()
time.sleep(5)

#extracting html table using bs4
table_element = driver.find_element(By.XPATH, "/html/body/center/table[1]")
table_html = table_element.get_attribute('outerHTML')
soup = BeautifulSoup(table_html, 'html.parser')

rows = soup.find_all('tr')
h = [th.text.strip() for th in rows[0].find_all('td')]
t_data = []

for row in rows[1:]:
    cells = row.find_all(('td','th'))
    row_data = {h[i]: cells[i].text.strip() for i in range(len(cells))}
    t_data.append(row_data)
print(t_data)

k=2
for j in table_data:
    #get last 10 digit of mobile number
    s= str(j['Seller Mobile'])
    sm= re.sub(r'\D', '', s)[-10:]
    b= str(j['Buyer Mobile'])
    bm= re.sub(r'\D', '', b)[-10:]
    n=2
    for m in t_data:
        if m['Mobile']==bm:
            buyer=driver.find_element(By.XPATH,f"/html/body/center/table[1]/tbody/tr{[n]}/td[1]/input")
            buyer.click()
            bn=m['Name']
        elif m['Mobile']==sm:
            seller=driver.find_element(By.XPATH,f"/html/body/center/table[1]/tbody/tr{[n]}/td[2]/input")
            seller.click()
            sn=m['Name']
        n+=1
    driver.implicitly_wait(10)
    pl_no=driver.find_element(By.XPATH,"/html/body/center/table[2]/tbody/tr[1]/td[2]/input")
    pl_no.click()
    pl_no.clear()
    pl_no.send_keys(j['Plot No'])
    driver.implicitly_wait(10)

    sqft=driver.find_element(By.XPATH,"/html/body/center/table[2]/tbody/tr[2]/td[2]/input")
    sqft.click()
    sqft.clear()
    sqft.send_keys(j['Sqft'])
    driver.implicitly_wait(10)

    sb=driver.find_element(By.XPATH,"/html/body/center/table[2]/tbody/tr[3]/td[2]/input")
    sb.click()
    driver.implicitly_wait(10)

    t=driver.find_element(By.ID,"TransNo")
    txt=t.text
    print(txt)
    sheet['G'+str(k)]=txt
    driver.back()
    k+=1

    user = 'abhaybadwaik.eidiko@gmail.com'
    password = 'cxfr lssm ayxq drjk'

    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login(user, password)

    msg = MIMEMultipart()
    msg['From'] = user
    msg['To'] = ''
    msg['Subject'] = f'Plot has booked Successfully-{txt}'
    # body of the email
    body = f"""
    Dear {sn} & {bn},<br><br>
    New Plot (Plot number:{j['Plot No']}) with no of sqft {j['Sqft']} has been booked successfully.<br><br>
    Here you can find Booking Details...<br><br>
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            table, th, td {{
            border: 1px solid black;
            border-collapse: collapse;
            padding: 5px;
        }}
        </style>
    </head>
    <body>
        <table style="width:100%">
            <tr>
                <td>Booking Number</td>
                <td>{txt}</td> 
            </tr>
            <tr>
                <td>Buyer Name</td>
                <td>{bn}</td>
            </tr>
            <tr>
                <td>Buyer Phone Number</td>
                <td>{bm}</td>
            </tr>
             <tr>
                <td>Seller Name</td>
                <td>{sn}</td>
            </tr>
             <tr>
                <td>Seller Phone Number</td>
                <td>{sm}</td>
            </tr>
        </table>
    </body>
    </html>
    Thanks.
    """
    msg.attach(MIMEText(body, 'html'))
    s.sendmail(user, "abhaybadwaik.eidiko@gmail.com", msg.as_string())

wb.save("/home/bandaru/Desktop/Selenium/input.xlsx")
driver.quit()
