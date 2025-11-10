import openpyxl
wb=openpyxl.load_workbook(r'/home/bandaru/Desktop/Selenium/txttask.xlsx')
ws=wb['Sheet1']
a=open('/home/bandaru/Desktop/Selenium/txt.txt','r')
j=2
for i in a:
    z=i.split(" ")
    print(z)
    y=i.split(":")
    print(y)
    print(y[-1])
    x=z[1].split(".")
    print(x)
    print(x[0])
    ws['A' + str(j)] = i[1:34]
    ws['B' + str(j)] = y[1][0:7]
    ws['C' + str(j)] = x[0]
    ws['D' + str(j)] = x[1]
    ws['E' + str(j)] = y[-1][0:7]
    ws['F' + str(j)] = y[-2][22:26]
    j+=1
wb.save(r'/home/bandaru/Desktop/Selenium/txttask.xlsx')