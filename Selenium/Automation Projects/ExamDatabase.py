import mysql.connector
mydb = mysql.connector.connect(
    host="127.0.0.1",
    user="root",
    password="786786",
    database="DB"
)
cursor = mydb.cursor()
# cursor.execute('CREATE TABLE sum(S_ID int, Tel int,Eng int,Maths int, Science int,Social int);')
import openpyxl
from openpyxl import load_workbook, Workbook
wb = load_workbook("/home/bandaru/Desktop/Selenium/Exam_Data.xlsx")
sheet = wb['Sheet1']
headers = [cell.value for cell in sheet[1]]
print(headers)
table_data = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
    data = {headers[i]: row[i].value for i in range(len(headers))}
    table_data.append(data)
print(table_data)
for j in table_data:
    a = "INSERT INTO exam(S_ID, Tel, Eng, Maths,Science,Social,total,average) VALUES (%s,%s,%s, %s, %s, %s,%s,%s)"
    val = (j['S_ID'], j['Tel'], j['Eng'], j['Maths'], j['Science'],j['Social'],j['total'],j['average'])

    cursor.execute(a, val)

# cursor.execute("SELECT DISTINCT * FROM exam_scores")
# data = cursor.fetchall()

mydb.commit()
wb.save("/home/bandaru/Desktop/Selenium/Exam_Data.xlsx")
wb.close()

